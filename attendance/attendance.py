import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_attendance_sheet(filename="january_attendance_2025.xlsx"):
    # Create a workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'January 2025'
    
    # Format headers
    headers = ['Name', 'Time in', 'Time out']
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 30  # Name
    ws.column_dimensions['B'].width = 15  # Time in
    ws.column_dimensions['C'].width = 15  # Time out
    
    # Save the workbook
    wb.save(filename)
    return filename

def add_daily_records(filename, date, records):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Add date header
    current_row = ws.max_row + 1
    date_cell = ws.cell(row=current_row, column=1)
    date_cell.value = date.strftime("%dth %B, %Y")
    date_cell.font = Font(color='4472C4')  # Blue color
    
    # Add records
    for name, time_in, time_out in records:
        current_row = ws.max_row + 1
        ws.cell(row=current_row, column=1, value=name)
        ws.cell(row=current_row, column=2, value=time_in)
        ws.cell(row=current_row, column=3, value=time_out)
        
        # Apply borders and center alignment to all cells in the row
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(left=Side(style='thin'), 
                               right=Side(style='thin'), 
                               top=Side(style='thin'), 
                               bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center')
    
    wb.save(filename)

# Example data - you can modify this
sample_employees = [
    "Nantongo Brendah",
    "Akello Esther",    
"Mirembe Jessica",
    "Namaganda Teddy",
    "Nantume Peninah",
    "Mwenyango Irene",
    "Atukwatse Ruth",
    "Ajio Harriet Peace",
    "Kiconco Topister",
    "Namayanja Veronica",
    "Namwase Racheal",
    "Bbosa Arnold Christopher",
    "Mpungu William Andrew",
    "Turyamureba Crispas",
    "Ahimbisibwe Sharon",
    "Ssemanda Geofrey",
    "Etuka Mirembe Sumaiya",
    "Akoth Sylivia",
    "Akoth Rose",
    "Mussira Yekosofat",
    "Musiimenta Morius",
    "Mukisa Simon",
    "Lunyoro Esther",
    "Ajoro Esther",
    "Nsambya Ian",
    "Lucia",
    "Costa"
]

if __name__ == "__main__":
    # Create the Excel file
    filename = create_attendance_sheet()
    
    # Add records for each day from January 8-31, 2025
    start_date = datetime(2025, 1, 8)
    end_date = datetime(2025, 1, 31)
    current_date = start_date
    
    while current_date <= end_date:
        # Sample records for demonstration
        # In real use, you would input actual attendance data
        daily_records = [
            (name, "", "") for name in sample_employees
        ]
        
        add_daily_records(filename, current_date, daily_records)
        current_date += timedelta(days=1)
    
    print(f"Excel file '{filename}' has been created successfully!")